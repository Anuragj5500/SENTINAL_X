"""
Reports API — Generate and download security reports in PDF/CSV/JSON formats.
Supports: Incident Report, Executive Summary, IOC Report, MITRE Report, Compliance.
"""
import os
import csv
import json
import io
from datetime import datetime, timezone, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse, FileResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_
from pydantic import BaseModel

from backend.database import get_db
from backend.models import Alert, Incident, Asset, ThreatIntelFeed, User, Report, Severity, AlertStatus
from backend.auth.rbac import get_current_user

router = APIRouter(prefix="/reports", tags=["Reports"])


# ─────────────────────────────── Schemas ─────────────────────────────────────

class ReportRequest(BaseModel):
    report_type: str  # incident, executive, ioc, mitre, compliance, asset
    format: str = "json"  # json, csv, pdf, excel
    days: int = 30
    title: Optional[str] = None


# ─────────────────────────────── Helpers ─────────────────────────────────────

async def _collect_executive_data(db: AsyncSession, days: int) -> dict:
    since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)

    total_alerts = (await db.execute(select(func.count(Alert.id)).where(Alert.created_at >= since))).scalar()
    critical_alerts = (await db.execute(
        select(func.count(Alert.id)).where(
            and_(Alert.created_at >= since, Alert.severity == Severity.critical)
        )
    )).scalar()
    total_incidents = (await db.execute(
        select(func.count(Incident.id)).where(Incident.created_at >= since)
    )).scalar()
    by_tactic = (await db.execute(
        select(Alert.mitre_tactic, func.count(Alert.id).label("count"))
        .where(and_(Alert.created_at >= since, Alert.mitre_tactic.isnot(None)))
        .group_by(Alert.mitre_tactic)
        .order_by(func.count(Alert.id).desc())
        .limit(10)
    )).all()
    top_attackers = (await db.execute(
        select(Alert.source_ip, func.count(Alert.id).label("count"))
        .where(and_(Alert.created_at >= since, Alert.source_ip.isnot(None)))
        .group_by(Alert.source_ip)
        .order_by(func.count(Alert.id).desc())
        .limit(10)
    )).all()

    return {
        "report_type": "Executive Summary",
        "period_days": days,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": {
            "total_alerts": total_alerts,
            "critical_alerts": critical_alerts,
            "total_incidents": total_incidents,
        },
        "top_tactics": [{"tactic": r[0], "count": r[1]} for r in by_tactic],
        "top_attackers": [{"ip": r[0], "count": r[1]} for r in top_attackers],
    }


async def _collect_incident_data(db: AsyncSession, days: int) -> dict:
    since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)
    result = await db.execute(
        select(Incident).where(Incident.created_at >= since).order_by(Incident.created_at.desc())
    )
    incidents = result.scalars().all()
    return {
        "report_type": "Incident Report",
        "period_days": days,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_incidents": len(incidents),
        "incidents": [
            {
                "id": i.id,
                "title": i.title,
                "severity": i.severity.value,
                "status": i.status.value,
                "priority": i.priority,
                "created_at": i.created_at.isoformat() if i.created_at else None,
                "resolved_at": i.resolved_at.isoformat() if i.resolved_at else None,
                "mitre_techniques": i.mitre_techniques,
                "tags": i.tags,
            }
            for i in incidents
        ],
    }


async def _collect_ioc_data(db: AsyncSession) -> dict:
    result = await db.execute(
        select(ThreatIntelFeed)
        .where(ThreatIntelFeed.is_active == True)
        .order_by(ThreatIntelFeed.confidence.desc())
    )
    iocs = result.scalars().all()
    return {
        "report_type": "IOC Report",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_iocs": len(iocs),
        "iocs": [
            {
                "ioc_value": i.ioc_value,
                "ioc_type": i.ioc_type,
                "threat_type": i.threat_type,
                "confidence": i.confidence,
                "source": i.source,
                "first_seen": i.first_seen.isoformat() if i.first_seen else None,
                "last_seen": i.last_seen.isoformat() if i.last_seen else None,
            }
            for i in iocs
        ],
    }


async def _collect_mitre_data(db: AsyncSession, days: int) -> dict:
    since = datetime.now(timezone.utc).replace(tzinfo=None) - timedelta(days=days)
    rows = (await db.execute(
        select(
            Alert.mitre_tactic,
            Alert.mitre_technique,
            func.count(Alert.id).label("count")
        )
        .where(and_(Alert.created_at >= since, Alert.mitre_technique.isnot(None)))
        .group_by(Alert.mitre_tactic, Alert.mitre_technique)
        .order_by(func.count(Alert.id).desc())
    )).all()

    return {
        "report_type": "MITRE ATT&CK Report",
        "period_days": days,
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "techniques": [
            {"tactic": r[0], "technique": r[1], "alert_count": r[2]}
            for r in rows
        ],
    }


async def _collect_asset_data(db: AsyncSession) -> dict:
    result = await db.execute(select(Asset).where(Asset.is_active == True))
    assets = result.scalars().all()
    return {
        "report_type": "Asset Report",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "total_assets": len(assets),
        "assets": [
            {
                "hostname": a.hostname,
                "ip_address": a.ip_address,
                "os_type": a.os_type,
                "criticality": a.criticality.value,
                "department": a.department,
                "antivirus_status": a.antivirus_status,
                "risk_score": a.risk_score,
                "agent_installed": a.agent_installed,
                "last_seen": a.last_seen.isoformat() if a.last_seen else None,
            }
            for a in assets
        ],
    }


def _to_csv(data: dict) -> str:
    """Convert report dict to CSV string."""
    output = io.StringIO()
    items_key = None
    for key in ["incidents", "iocs", "techniques", "assets", "top_tactics", "top_attackers"]:
        if key in data:
            items_key = key
            break

    if items_key and data[items_key]:
        writer = csv.DictWriter(output, fieldnames=data[items_key][0].keys())
        writer.writeheader()
        writer.writerows(data[items_key])
    else:
        output.write(json.dumps(data, indent=2))

    return output.getvalue()


# ─────────────────────────────── Routes ──────────────────────────────────────

@router.get("")
async def list_reports(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(
        select(Report).order_by(Report.created_at.desc()).limit(50)
    )
    reports = result.scalars().all()
    return [
        {
            "id": r.id,
            "title": r.title,
            "report_type": r.report_type,
            "status": r.status,
            "created_at": r.created_at,
        }
        for r in reports
    ]


@router.post("/generate")
async def generate_report(
    req: ReportRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Generate a report and return it immediately as JSON, CSV, PDF, or Excel."""
    report_type = req.report_type.lower()

    if report_type == "executive":
        data = await _collect_executive_data(db, req.days)
    elif report_type == "incident":
        data = await _collect_incident_data(db, req.days)
    elif report_type == "ioc":
        data = await _collect_ioc_data(db)
    elif report_type == "mitre":
        data = await _collect_mitre_data(db, req.days)
    elif report_type == "asset":
        data = await _collect_asset_data(db)
    elif report_type == "compliance":
        from backend.compliance import evaluate_all_frameworks
        data = await evaluate_all_frameworks(db, req.days)
        data["report_type"] = "Compliance Report"
    else:
        raise HTTPException(400, f"Unknown report type: {req.report_type}. Use: executive, incident, ioc, mitre, asset, compliance")

    # Log the report generation
    title = req.title or f"{data['report_type']} — {datetime.now(timezone.utc).strftime('%Y-%m-%d')}"
    report_record = Report(
        title=title,
        report_type=report_type,
        generated_by=current_user.id,
        params={"days": req.days, "format": req.format},
        status="completed",
    )
    db.add(report_record)
    await db.flush()

    if req.format == "csv":
        csv_content = _to_csv(data)
        return StreamingResponse(
            iter([csv_content]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="sentinelx_{report_type}.csv"'},
        )

    if req.format == "pdf":
        pdf_bytes = _generate_pdf(data, title)
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="sentinelx_{report_type}.pdf"'},
        )

    if req.format == "excel":
        excel_bytes = _generate_excel(data, title)
        return StreamingResponse(
            iter([excel_bytes]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="sentinelx_{report_type}.xlsx"'},
        )

    return data


# ───────────────────────────── PDF Generation ──────────────────────────────

def _generate_pdf(data: dict, title: str) -> bytes:
    """Generate a PDF report using reportlab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.colors import HexColor
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import inch

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'],
                                  fontSize=22, textColor=HexColor('#00d4ff'), spaceAfter=20)
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'],
                                    textColor=HexColor('#1a73e8'), spaceAfter=10)
    body_style = styles['BodyText']

    elements = []
    elements.append(Paragraph(f"🛡️ SentinelX — {title}", title_style))
    elements.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}", body_style))
    elements.append(Spacer(1, 20))

    report_type = data.get('report_type', 'Report')
    elements.append(Paragraph(f"Report Type: {report_type}", heading_style))
    elements.append(Spacer(1, 10))

    # Summary section
    if 'summary' in data:
        elements.append(Paragraph("Summary", heading_style))
        for key, val in data['summary'].items():
            elements.append(Paragraph(f"<b>{key.replace('_', ' ').title()}:</b> {val}", body_style))
        elements.append(Spacer(1, 15))

    # Table data
    items_key = None
    for key in ['incidents', 'iocs', 'techniques', 'assets', 'top_tactics', 'top_attackers', 'controls']:
        if key in data and data[key]:
            items_key = key
            break

    if items_key and data[items_key]:
        elements.append(Paragraph(items_key.replace('_', ' ').title(), heading_style))
        rows = data[items_key]
        if rows:
            headers = list(rows[0].keys())[:6]  # Limit columns for readability
            table_data = [headers]
            for row in rows[:50]:  # Limit rows
                table_data.append([str(row.get(h, ''))[:40] for h in headers])

            t = Table(table_data, repeatRows=1)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), HexColor('#1a237e')),
                ('TEXTCOLOR', (0, 0), (-1, 0), HexColor('#ffffff')),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [HexColor('#f5f5f5'), HexColor('#ffffff')]),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]))
            elements.append(t)

    elements.append(Spacer(1, 30))
    elements.append(Paragraph("<i>Report generated by SentinelX SIEM Platform</i>", body_style))

    doc.build(elements)
    return buffer.getvalue()


def _generate_excel(data: dict, title: str) -> bytes:
    """Generate an Excel report using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
    title_font = Font(bold=True, size=16, color="0066CC")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"SentinelX — {title}"
    ws['A1'].font = title_font
    ws['A2'] = f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws['A2'].font = Font(italic=True, color="666666")

    row = 4

    # Summary
    if 'summary' in data:
        ws.cell(row=row, column=1, value="Summary").font = Font(bold=True, size=13)
        row += 1
        for key, val in data['summary'].items():
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            ws.cell(row=row, column=2, value=str(val))
            row += 1
        row += 1

    # Data table
    items_key = None
    for key in ['incidents', 'iocs', 'techniques', 'assets', 'top_tactics', 'controls']:
        if key in data and data[key]:
            items_key = key
            break

    if items_key and data[items_key]:
        rows_data = data[items_key]
        headers = list(rows_data[0].keys())

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header.replace('_', ' ').title())
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        row += 1

        # Write data
        for item in rows_data:
            for col, header in enumerate(headers, 1):
                val = item.get(header, '')
                if isinstance(val, (list, dict)):
                    val = json.dumps(val)[:100]
                cell = ws.cell(row=row, column=col, value=str(val)[:200])
                cell.border = thin_border
            row += 1

        # Auto-width columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[ws.cell(row=4, column=col).column_letter].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.get("/templates")
async def report_templates(current_user: User = Depends(get_current_user)):
    """Return available report templates."""
    return [
        {
            "id": "executive",
            "name": "Executive Summary",
            "description": "High-level security posture overview for leadership",
            "default_days": 30,
            "formats": ["json", "csv", "pdf", "excel"],
        },
        {
            "id": "incident",
            "name": "Incident Report",
            "description": "Detailed log of all security incidents in the period",
            "default_days": 30,
            "formats": ["json", "csv", "pdf", "excel"],
        },
        {
            "id": "mitre",
            "name": "MITRE ATT&CK Report",
            "description": "ATT&CK technique coverage and attack frequency heatmap",
            "default_days": 30,
            "formats": ["json", "csv", "pdf", "excel"],
        },
        {
            "id": "ioc",
            "name": "IOC Report",
            "description": "All active Indicators of Compromise in the threat intel feed",
            "default_days": None,
            "formats": ["json", "csv", "pdf", "excel"],
        },
        {
            "id": "asset",
            "name": "Asset Inventory Report",
            "description": "Complete asset inventory with risk scores and agent status",
            "default_days": None,
            "formats": ["json", "csv", "pdf", "excel"],
        },
        {
            "id": "compliance",
            "name": "Compliance Report",
            "description": "PCI DSS, SOC2, ISO27001, HIPAA, GDPR compliance scores and findings",
            "default_days": 30,
            "formats": ["json", "csv", "pdf", "excel"],
        },
    ]
